from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework.response import Response 
from django.http import HttpResponse
from rest_framework import status
import csv
from openpyxl import load_workbook
from .models import grade,section,studentdetails,teacherdetails,attendance
from django.utils import timezone
from .serializers import GradeSerializer, SectionSerializer,StudentDetailSerializer,TeacherDetailSerializer,attendanceSerializer
from users.models import users
class GradeView(APIView):
    def post(self,request):
        try:
            serializer = GradeSerializer(data = request.data)
            if serializer.is_valid(): 
                serializer.save()           
                return Response({
                    'message':"Grade create successfully",
                    'status':200
                },status = status.HTTP_200_OK)
            else:
                return Response({
                    'message':"Error while creating grade",
                    'status':500
                },status = status.HTTP_400_BAD_REQUEST)

        except Exception as error:
            return Response({
                'message':f'{str(error)}',
                'status':500
            },status = status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    def get(self,request):
        try:
            gradeData = grade.objects.all()
            if gradeData.exists:
                serializer = GradeSerializer(gradeData,many=True)
                return Response({
                    'message':"Grade data retrieved successfully",
                    'status':201,
                    'result':serializer.data
                },status = status.HTTP_200_OK)
                
            else:
                return Response({
                    'message':"No data found",
                    'status':201
                },status = status.HTTP_200_OK)

        except Exception as error:
            return Response({
                'message':f'{str(error)}',
                'status':500
            },status = status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    def put(self,request):
        try:
            grade_id = request.query_params.get('grade_id')
            gradeData = grade.objects.get(grade_id = grade_id)
            if gradeData:
                serializer = GradeSerializer(gradeData,data = request.data)
                if serializer.is_valid(): 
                    serializer.save()           
                    return Response({
                        'message':"Grade updated successfully",
                        'status':200
                    },status = status.HTTP_200_OK)
                else:
                    return Response({
                        'message':"Error while getting grade",
                        'status':400
                    },status = status.HTTP_400_BAD_REQUEST)
            else:
                return Response({
                    'message':"No grade found to update",
                    'status':200
                },status = status.HTTP_200_OK)

        except Exception as error:
            return Response({
                'message':f'{str(error)}',
                'status':500
            },status = status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self,request):
        try:
            grade_id = request.query_params.get('grade_id')
            gradeData = grade.objects.filter(grade_id = grade_id)
            if gradeData:
                gradeData.delete()
                return Response({
                    'message':"Grade deleted successfully",
                    'status':201
                },status = status.HTTP_200_OK)
            else:
                return Response({
                    'message':"No grade found to delete",
                    'status':201
                },status = status.HTTP_200_OK)

        except Exception as error:
            return Response({
                'message':f'{str(error)}',
                'status':500
            },status = status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    

        
    
class SectionView(APIView):
    def post(self,request):
        try:
            serializer = SectionSerializer(data = request.data)
            if serializer.is_valid(): 
                serializer.save()           
                return Response({
                    'message':"Section create successfully",
                    'status':200
                },status = status.HTTP_200_OK)
            else:
                return Response({
                    'message':"Error while creating section",
                    'status':500
                },status = status.HTTP_400_BAD_REQUEST)

        except Exception as error:
            return Response({
                'message':f'{str(error)}',
                'status':500
            },status = status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    def get(self,request):
        try:
            sectionData = section.objects.all()
            if sectionData.exists:
                serializer = SectionSerializer(sectionData,many=True)
                return Response({
                    'message':"Section data retrieved successfully",
                    'status':201,
                    'result':serializer.data
                },status = status.HTTP_200_OK)
                
            else:
                return Response({
                    'message':"No data found",
                    'status':201
                },status = status.HTTP_200_OK)

        except Exception as error:
            return Response({
                'message':f'{str(error)}',
                'status':500
            },status = status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    def put(self,request):
        try:
            section_id = request.query_params.get('section_id')
            sectionData = section.objects.get(section_id = section_id)
            if sectionData:
                serializer = SectionSerializer(sectionData,data = request.data)
                if serializer.is_valid(): 
                    serializer.save()           
                    return Response({
                        'message':"Section updated successfully",
                        'status':200
                    },status = status.HTTP_200_OK)
                else:
                    return Response({
                        'message':"Error while getting section",
                        'status':400
                    },status = status.HTTP_400_BAD_REQUEST)
            else:
                return Response({
                    'message':"No section found to update",
                    'status':200
                },status = status.HTTP_200_OK)

        except Exception as error:
            return Response({
                'message':f'{str(error)}',
                'status':500
            },status = status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self,request):
        try:
            section_id = request.query_params.get('section_id')
            sectionData = section.objects.filter(section_id = section_id)
            if sectionData:
                sectionData.delete()
                return Response({
                    'message':"Section deleted successfully",
                    'status':201
                },status = status.HTTP_200_OK)
            else:
                return Response({
                    'message':"No section found to delete",
                    'status':201
                },status = status.HTTP_200_OK)

        except Exception as error:
            return Response({
                'message':f'{str(error)}',
                'status':500
            },status = status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    
@api_view(['GET'])
def generateAttendanceCsv(request):
    try:
        fromDate = request.query_params.get('from_date')
        toDate = request.query_params.get('to_date')
        grade = request.query_params.get('grade')
        section = request.query_params.get('section')
        teacher_id = request.query_params.get('teacher_id')
        if fromDate and toDate:
            getAttendanceData = attendance.objects.filter(grade_name = grade , section_name = section ,
                                                    teacher_id = teacher_id,
                                                    updated_at__range = (fromDate,toDate) ).all()
        else:
            getAttendanceData = attendance.objects.filter(grade_name = grade , section_name = section ,
                                                    teacher_id = teacher_id)
        if not getAttendanceData:
            return Response({
            'message':'No attendance Found',
            'status':200,
        },status=status.HTTP_200_OK)

        excelData = []
        for attendances in getAttendanceData:
            excelData.append([
                attendances.grade_name,
                attendances.section_name,
                attendances.student_name,
                attendances.attendance_status,
                attendances.remark
            ])
        
        response = HttpResponse(
            content_type='text/csv',
            headers={'Content-Disposition': 'attachment; filename="your_data.csv"'},
        )
        # Create a CSV writer object.
        writer = csv.writer(response)

        # Write the header row (e.g., column names).
        writer.writerow(['Grade', 'Section', 'Student','Attendance Status','Remark'])

        for row in excelData:
            writer.writerow(row)
        return response

    except Exception as error:
        return Response({
            'message':f'{str(error)}',
            'status':500,
        },status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST']) 
def uploadAttendanceExcel(request):
    try:
        teacher_id = request.POST.get('teacher_id')
        file = request.FILES.get('file')

        if not teacher_id or not file:
            return Response({
                'message': 'Missing teacher_id or file',
                'status': 400
            }, status=status.HTTP_400_BAD_REQUEST)

        # ✅ Read using openpyxl
        wb = load_workbook(file)
        ws = wb.active  # first sheet

        # Skip the header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            student_name, grade_val, section_name, attendance_status, remark = row
            if not student_name:
                continue

            grade_obj = grade.objects.filter(grade_name=str(grade_val).strip()).first()
            section_obj = section.objects.filter(section_name=str(section_name).strip()).first()

            # Parse name
            name_parts = str(student_name).strip().split(" ", 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ""

            user_obj = users.objects.filter(
                first_name=first_name,
                last_name=last_name
            ).first()

            student_obj = None
            if user_obj:
                student_obj = studentdetails.objects.filter(
                    student_id=user_obj.user_id,
                    grade_name=str(grade_val).strip(),
                    section_name=str(section_name).strip()
                ).first()

            if not all([grade_obj, section_obj, student_obj]):
                print(f"Skipping row: {row}")
                continue

            attendance.objects.create(
                teacher_id=teacher_id,
                student_id=student_obj.student_id,
                student_name=student_name,
                grade_id=grade_obj.grade_id,
                grade_name=grade_obj.grade_name,
                section_id=section_obj.section_id,
                section_name=section_obj.section_name,
                attendance_status=attendance_status,
                remark=remark,
                created_at=timezone.now(),
                updated_at=timezone.now()
            )


        return Response({
            'message': 'Attendance uploaded successfully',
            'status': 200
        }, status=status.HTTP_200_OK)

    except Exception as error:
        return Response({
            'message': f'{str(error)}',
            'status': 500
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
